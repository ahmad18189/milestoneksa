import frappe, os, re
from frappe.utils import getdate, cstr

# ---------------------- Utils ----------------------

def _safe_float(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def _norm(s):
    """Normalize labels/headers for fuzzy matching."""
    if s is None:
        return ""
    s = cstr(s)
    s = s.strip().lower()
    # drop diacritics-like chars and punctuation/spaces
    s = re.sub(r"[^\w]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _read_xlsx(path, sheet_name="Sheet1"):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    try:
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet_name)
        df.columns = [c.strip() for c in df.columns]
        df = df.where(pd.notnull(df), None)
        return df.to_dict(orient="records"), list(df.columns)
    except ImportError:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise Exception(f"Sheet '{sheet_name}' not found; sheets: {wb.sheetnames}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [c.strip() if isinstance(c, str) else c for c in (rows[0] or [])]
        out = []
        for r in rows[1:]:
            rec = {}
            for i, h in enumerate(headers):
                if h:
                    rec[h] = r[i] if i < len(r) else None
            out.append(rec)
        return out, headers

# ---------------------- Build mapping from JSON/meta ----------------------

def _build_label_map_for_parent(fields_json=None):
    """
    Build a mapping from normalized label/fieldname -> fieldname for parent (Journal Entry).
    Uses your Customize Form JSON if provided; otherwise falls back to frappe.get_meta.
    """
    mapping = {}
    known_parent = {}  # friendly fallbacks for common headers
    # Friendly fallbacks
    known_parent.update({
        _norm("Entry Type"): "voucher_type",
        _norm("Series"): "naming_series",
        _norm("Company"): "company",
        _norm("Posting Date"): "posting_date",
        _norm("Title"): "title",
        _norm("Old Journal Number"): "custom_old_journal_number",
        _norm("Reference Number"): "cheque_no",
        _norm("Reference Date"): "cheque_date",
        _norm("User Remark"): "user_remark",
        _norm("Remark"): "remark",
        _norm("Pay To / Recd From"): "pay_to_recd_from",
        _norm("Mode of Payment"): "mode_of_payment",
        _norm("Due Date"): "due_date",
        _norm("Bill No"): "bill_no",
        _norm("Bill Date"): "bill_date",
        _norm("Is Opening"): "is_opening",
    })

    if fields_json:
        for f in fields_json:
            if f.get("doctype") not in ("Customize Form Field", "DocField", None):
                continue
            fn = f.get("fieldname")
            lb = f.get("label")
            if fn:
                mapping[_norm(fn)] = fn
            if lb:
                mapping[_norm(lb)] = fn
    else:
        meta = frappe.get_meta("Journal Entry")
        for df in meta.fields:
            if df.fieldname:
                mapping[_norm(df.fieldname)] = df.fieldname
            if df.label:
                mapping[_norm(df.label)] = df.fieldname

    # last: add the friendly fallbacks
    mapping.update(known_parent)
    return mapping

def _build_label_map_for_child():
    """
    Build mapping for Journal Entry Account (child table).
    """
    mapping = {}
    fallbacks = {
        _norm("Account"): "account",
        _norm("Debit"): "debit",
        _norm("Credit"): "credit",
        _norm("Cost Center"): "cost_center",
        _norm("Party Type"): "party_type",
        _norm("Party"): "party",
        _norm("User Remark"): "user_remark",
        _norm("Reference Type"): "reference_type",
        _norm("Reference Name"): "reference_name",
    }

    try:
        meta = frappe.get_meta("Journal Entry Account")
        for df in meta.fields:
            if df.fieldname:
                mapping[_norm(df.fieldname)] = df.fieldname
            if df.label:
                mapping[_norm(df.label)] = df.fieldname
    except Exception:
        pass

    mapping.update(fallbacks)
    return mapping

def _split_child_header(h):
    """
    If a header is like 'Account (Accounting Entries)' returns ('Account', True).
    Otherwise returns (h, False).
    """
    m = re.match(r"^(.*)\((.*?)\)\s*$", h or "")
    if not m:
        return h, False
    base = m.group(1).strip()
    # tag = m.group(2).strip()  # 'Accounting Entries'
    return base, True

# ---------------------- Row mapping ----------------------

def _parent_values_from_row(row, parent_map, headers):
    data = {}
    for h in headers:
        base = h
        is_child_hint = False
        if "(" in h and ")" in h:
            base, is_child_hint = _split_child_header(h)
            if is_child_hint:
                continue  # belongs to child
        key = parent_map.get(_norm(base))
        if key:
            data[key] = row.get(h)
    return data

def _child_values_from_row(row, child_map, headers):
    """
    Build one child row from the columns marked as '(Accounting Entries)'.
    """
    c = {}
    for h in headers:
        base = h
        is_child_hint = False
        if "(" in h and ")" in h:
            base, is_child_hint = _split_child_header(h)
        else:
            continue  # we only take explicitly marked child columns
        if not is_child_hint:
            continue
        key = child_map.get(_norm(base))
        if not key:
            continue
        v = row.get(h)
        # number fields:
        if key in ("debit", "credit"):
            v = _safe_float(v)
        c[key] = v

    # Skip empty lines
    if not any([c.get("account"), _safe_float(c.get("debit")), _safe_float(c.get("credit"))]):
        return None
    return c

# ---------------------- Import main ----------------------

def import_journal_entries(
    xlsx_path,
    sheet_name="Sheet1",
    fields_json=None,         # paste your Customize Form JSON array here or pass from caller
    submit=True,
    dry_run=False,
    group_key_candidates=("Old Journal Number", "custom_old_journal_number", "Old Journal #")
):
    rows, headers = _read_xlsx(xlsx_path, sheet_name=sheet_name)
    if not rows:
        print("No rows found. Aborting.")
        return

    parent_map = _build_label_map_for_parent(fields_json)
    child_map = _build_label_map_for_child()

    # detect group key
    group_key_header = None
    norm_headers = {_norm(h): h for h in headers}
    for cand in group_key_candidates:
        nh = _norm(cand)
        if nh in norm_headers:
            group_key_header = norm_headers[nh]
            break
    if not group_key_header:
        # if not found, try if parent_map resolves to the custom field and header exists
        target_fn = "custom_old_journal_number"
        for h in headers:
            if parent_map.get(_norm(h)) == target_fn:
                group_key_header = h
                break
    if not group_key_header:
        raise Exception("Group key not found. Add a column 'Old Journal Number' or 'custom_old_journal_number'.")

    # group
    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        key = r.get(group_key_header) or f"ROW-{len(groups)+1}"
        groups[key].append(r)

    keys = list(groups.keys())
    total = len(keys)
    print(f"Found {len(rows)} rows -> {total} journal groups by '{group_key_header}'.")
    print(f"submit={submit} dry_run={dry_run}")
    print("-"*80)

    # Report
    rep = []
    ok, fail = 0, 0

    for i, gkey in enumerate(keys, 1):
        grows = groups[gkey]
        parent_vals = _parent_values_from_row(grows[0], parent_map, headers)

        je = frappe.new_doc("Journal Entry")
        for k,v in parent_vals.items():
            if k == "posting_date" and v:
                try:
                    v = getdate(v)
                except Exception:
                    pass
            setattr(je, k, v)

        # If the custom field exists, store the Old Journal Number too
        try:
            if hasattr(je, "custom_old_journal_number") and not getattr(je, "custom_old_journal_number", None):
                setattr(je, "custom_old_journal_number", gkey)
        except Exception:
            pass

        je.set("accounts", [])
        for r in grows:
            ch = _child_values_from_row(r, child_map, headers)
            if ch:
                je.append("accounts", ch)

        # Totals check
        tdr = sum(_safe_float(c.debit) for c in je.accounts)
        tcr = sum(_safe_float(c.credit) for c in je.accounts)
        if round(tdr, 2) != round(tcr, 2):
            msg = f"Dr/Cr mismatch in group '{gkey}' (Dr={tdr}, Cr={tcr})"
            print(f"[{i}/{total}] ❌ {msg}")
            rep.append({"Group": gkey, "Status": "Failed", "Name": "", "Rows": len(je.accounts), "Error": msg})
            fail += 1
            continue

        try:
            if dry_run:
                je.flags.ignore_permissions = True
                je.set_missing_values()
                je.validate()
                print(f"[{i}/{total}] ✅ DRY-RUN OK {gkey} (rows={len(je.accounts)})")
                rep.append({"Group": gkey, "Status": "Dry-Run OK", "Name": "", "Rows": len(je.accounts), "Error": ""})
            else:
                je.insert(ignore_permissions=True)
                if submit:
                    je.submit()
                frappe.db.commit()
                print(f"[{i}/{total}] ✅ Imported {je.name} (group={gkey}, rows={len(je.accounts)})")
                rep.append({"Group": gkey, "Status": "Imported", "Name": je.name, "Rows": len(je.accounts), "Error": ""})
                ok += 1
        except Exception as ex:
            frappe.db.rollback()
            err = cstr(ex)
            frappe.log_error(err, title=f"JE Import Error [{gkey}]")
            print(f"[{i}/{total}] ❌ ERROR group={gkey}: {err}")
            rep.append({"Group": gkey, "Status": "Failed", "Name": "", "Rows": len(je.accounts), "Error": err})
            fail += 1

    # write CSV report
    try:
        import csv
        site_path = frappe.get_site_path("private", "files")
        os.makedirs(site_path, exist_ok=True)
        out_csv = os.path.join(site_path, "je_import_report.csv")
        with open(out_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["Group", "Status", "Name", "Rows", "Error"])
            w.writeheader()
            for r in rep:
                w.writerow(r)
        print("-"*80)
        print(f"Report written to: {out_csv}")
    except Exception as ex:
        print(f"Could not write report CSV: {ex}")

    print("-"*80)
    print(f"Done. Imported: {ok}, Failed: {fail}, Total groups: {total}")


# apps/milestoneksa/milestoneksa/jv_import.py
import frappe
from frappe.core.doctype.data_import.importer import Importer

def run(doctype, path, import_type="Insert New Records", submit=False, console=True): 
    """
    import_type: "Insert New Records" or "Update Existing Records"
    submit: submit after import if DocType is submittable (e.g., Journal Entry)
    """
    # Build a transient Data Import doc to carry options
    di = frappe.new_doc("Data Import")
    di.import_type = import_type
    di.submit_after_import = 1 if submit else 0
    di.mute_emails = 1
    di.reference_doctype = doctype
    di.insert()

    imp = Importer(doctype=doctype, data_import=di, file_path=path, console=console)
    logs = imp.import_data()  # returns a list of log dicts when console=True
    # Optional: print a minimal summary for CLI usage
    success = sum(1 for l in logs if l.get("success"))
    fail = sum(1 for l in logs if not l.get("success"))
    print(f"[{doctype}] import finished: {success} success, {fail} failed")


import os
import frappe
from frappe.core.doctype.data_import.importer import Importer

def run_static():
    """
    One-shot Journal Entry import from XLSX with print debugging.
    """

    # ---- STATIC CONFIG ----
    doctype = "Journal Entry"
    abs_path = "/home/erp/frappe-bench/sites/erp.olivegardensksa.com/private/files/journal_entry_full_import_ids_corrected6fe544.xlsx"
    import_type = "Insert New Records"   # or "Update Existing Records"
    auto_submit = True
    # -----------------------

    print(f"[INFO] Starting static import for {doctype}")
    print(f"[INFO] File path: {abs_path}")
    if not os.path.exists(abs_path):
        raise Exception(f"File not found: {abs_path}")

    # 1) Ensure File doc
    guessed_url = "/private/files/" + os.path.basename(abs_path)
    print(f"[DEBUG] Checking if File already exists at {guessed_url}")
    existing_name = frappe.db.get_value("File", {"file_url": guessed_url})
    if existing_name:
        file_url = guessed_url
        print(f"[INFO] Using existing File: {file_url}")
    else:
        print("[INFO] Creating new File doc...")
        with open(abs_path, "rb") as f:
            content = f.read()
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": os.path.basename(abs_path),
            "content": content,
            "is_private": 1
        }).insert(ignore_permissions=True)
        file_url = file_doc.file_url
        print(f"[INFO] Created new File: {file_url}")

    # 2) Data Import doc
    print("[INFO] Creating Data Import document...")
    di = frappe.new_doc("Data Import")
    di.reference_doctype = doctype
    di.import_type = import_type
    di.submit_after_import = 1 if auto_submit else 0
    di.mute_emails = 1
    di.import_file = file_url
    di.insert(ignore_permissions=True)
    print(f"[INFO] Data Import created: {di.name}")

    # 3) Run Importer
    print("[INFO] Running Importer...")
    imp = Importer(doctype=doctype, data_import=di, console=False)
    logs = imp.import_data() or []
    print(f"[INFO] Import finished with {len(logs)} log entries")

    # 4) Summarize
    success = sum(1 for l in logs if l.get("success"))
    failed = sum(1 for l in logs if not l.get("success"))
    status = frappe.db.get_value("Data Import", di.name, "status")

    print(f"[RESULT] Status: {status}")
    print(f"[RESULT] Success: {success}, Failed: {failed}")

    return {
        "data_import": di.name,
        "doctype": doctype,
        "file_url": file_url,
        "status": status,
        "success": success,
        "failed": failed,
    }
